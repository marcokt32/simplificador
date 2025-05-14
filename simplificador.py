import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

def processar_excel(caminho_entrada, caminho_saida, tipo_base):
    # Define a base correta
    if tipo_base == "tc":
        caminho_base = "base_tc.xlsx"
    elif tipo_base == "wso":
        caminho_base = "base_wso.xlsx"
    else:
        raise ValueError("Tipo de base inválido: deve ser 'tc' ou 'wso'")

    df = pd.read_excel(caminho_entrada, sheet_name=0, skiprows=4)
    df.columns = df.iloc[0]
    df = df[1:]
    df = df.loc[:, ~df.columns.duplicated()]
    df = df.loc[:, df.columns.notna()]

    print("🔍 Colunas detectadas:", list(df.columns))

    colunas_desejadas = [
        "Unidades", "# de anúncio",
        "Título do anúncio", "Variação", "Comprador"
    ]

    colunas_existentes = [col for col in colunas_desejadas if col in df.columns]
    if not colunas_existentes:
        raise ValueError("❌ Nenhuma das colunas esperadas foi encontrada.")

    # Usa a base selecionada
    base_df = pd.read_excel(caminho_base)
    if '# de anúncio' not in base_df.columns or 'Nome do anúncio' not in base_df.columns:
        raise ValueError("❌ A base deve conter '# de anúncio' e 'Nome do anúncio'.")

    base_dict = dict(zip(base_df['# de anúncio'], base_df['Nome do anúncio']))
    df['Título do anúncio'] = df['# de anúncio'].map(base_dict)
    df_simplificado = df[colunas_existentes]
    df_simplificado.to_excel(caminho_saida, index=False)

    wb = openpyxl.load_workbook(caminho_saida)
    ws = wb.active
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 22.5
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
    wb.save(caminho_saida)